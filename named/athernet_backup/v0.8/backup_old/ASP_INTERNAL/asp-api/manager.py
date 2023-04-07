from flask import Blueprint, g, request, jsonify
import sqlite3
from tokandgen import checkTokenValidHandler
from misc import dict_factory, check_bus_or_user, ReArrangeData
from pagerrors import page_not_found, internal_error, forbbid_any
import time
import openpyxl
import re
from openpyxl.styles import Alignment, Font

ASP_FOLDER = "/home/kj/Desktop/.d/ASP_INTERNAL"
OFFER_DIRECTORIES = ASP_FOLDER+"/PRICE_OFFERS/"
manager_api = Blueprint('manager_api', __name__)




'''
This Function is used to fetch all the pending items for orders and quotations
'''
@manager_api.route('/pos/handler', methods=['GET'])
def api_serv_fetch_pending_items(key):

    g.key = key
    if not checkTokenValidHandler(key):
        return forbbid_any(request) 
  
    point_guy_id = request.args.get('id')
    qorp = request.args.get('q')
    penordon = request.args.get('d')    

    query = "SELECT irequestid,concode,quorpo,timestamp,done FROM busreq WHERE pointguyid=? AND done=? AND quorpo=?;"

    if (point_guy_id.isdigit() and qorp.isdigit() and penordon.isdigit()):
        to_filter = [point_guy_id, "0" if penordon == "0" else "1", 'Purchase' if qorp == '1' else 'Quote']
    else:
        return forbbid_any(request)

    base_conn = sqlite3.connect(ASP_FOLDER+'/databases/base.db')
    base_conn.row_factory = dict_factory
    base_cur = base_conn.cursor()
    lite_answer = base_cur.execute(query, to_filter).fetchall()

    if len(lite_answer) == 0:
        return forbbid_any(request)

    results = []
    for i in range(len(lite_answer)):
        if (check_bus_or_user("irequestid", lite_answer[i]['irequestid']) == 'usr'):
            query2 = """SELECT contactperson,phonenumber FROM users WHERE irequestid=?;"""
        elif (check_bus_or_user("irequestid", lite_answer[i]['irequestid']) == 'bus'):
            query2 = """SELECT name,phonenumber,contactperson FROM businesses WHERE irequestid=?;"""
        else:
            continue
        part_res = base_cur.execute(query2,[lite_answer[i]['irequestid']]).fetchall()
        part_res[0]['concode'] =  lite_answer[i]['concode']
        part_res[0]['quorpo'] = lite_answer[i]['quorpo']
        part_res[0]['timestamp'] = lite_answer[i]['timestamp']
        results.append(part_res)

    base_cur.close()
    base_conn.close()
    return jsonify(results)


'''
This function is used to fetch the stock avaliable
'''
@manager_api.route('/fetch/stock/', methods=['GET'])
def api_serv_fetch_stock(key):
    g.key = key
    if not checkTokenValidHandler(key):
        return forbbid_any(request) 
   
    temp_query = '%'+request.args.get('q')+'%'


    query = "SELECT * FROM oldstock WHERE partnumber like ?;"
    to_filter = [temp_query]
    results = []
    print(query, to_filter)
    parts_conn = sqlite3.connect(ASP_FOLDER+'/databases/parts.db')
    parts_conn.row_factory = dict_factory
    part_cur = parts_conn.cursor()
    results = part_cur.execute(query,to_filter).fetchall()
    print(results)
    parts_conn.close()
    if len(results) < 1:
        return """{} results""".format("0")
    else:    
        return jsonify(results)



'''
This function is used to send all the done orders and quotations to their respective handler
'''
@manager_api.route('/done/handler/items', methods=['GET'])
def api_serv_fetch_done_items(key):

    g.key = key
    if not checkTokenValidHandler(key):
        return forbbid_any(request) 

    # get the handler ID and whether quotations or purchase is being fetched
    # from the URL request arguments    
    cncd = request.args.get('c')
    qorp = request.args.get('q')
    to_filter = []

    #firstly query all the requests for     

    if (cncd.isdigit()):
        to_filter.append(cncd)
    else:
        return page_not_found(request)

    if not qorp.isdigit():
        return page_not_found(request)
    else:    
        if qorp == "1":
            query = "SELECT * FROM orders WHERE concod=?;"
        elif qorp == "2":
            query = "SELECT * FROM quotations WHERE concod=?;"
        else:
            return page_not_found(request) 
        
        results = []

        conn1 = sqlite3.connect(ASP_FOLDER+'/databases/action.db')
        conn1.row_factory = dict_factory
        cur1 = conn1.cursor()
        results = cur1.execute(query,to_filter).fetchall()

        conn1.close()
        if len(results) < 1:
            return page_not_found(request)
        else:    
            return jsonify(results)



'''
This function is used to fetch all pending orders and quotations ITEMS
'''
@manager_api.route('/pos/all', methods=['GET'])
def api_serv_fetch_pending_ids(key):
    g.key = key
    if not checkTokenValidHandler(key):
        return forbbid_any(request)
    temp_concode = request.args.get('c').strip()
    temp_pointguyid = request.args.get('id').strip()
    if not (temp_concode.isdigit() and temp_pointguyid.isdigit()):
        return page_not_found(request)
    else:
        lite_answer = []
        query = 'SELECT * FROM requests WHERE concod=? AND pointguyid=?;'
        to_filter = [temp_concode, temp_pointguyid]
        conn = sqlite3.connect(ASP_FOLDER+'/databases/action.db')
        conn.row_factory = dict_factory
        cur = conn.cursor()
        lite_answer = cur.execute(query, to_filter).fetchall()
        conn.close()
        results = []
        for i in lite_answer:
            try:
                to_add = priceit(i['partnumber'], i['designation'], i['brand'], i['quantity'])[0]
                i['sprice'] = to_add['price']
                i['saval'] = to_add['quantity']                    
            except:
                pass    
        print(lite_answer)
        return jsonify(lite_answer)


'''
This Function is used to distubiute the main requests into either a quotations or an orders.
And it will be deleted from the requests table
'''
@manager_api.route('/submit', methods=['POST'])
def api_serv_exec_form(key):
    g.key = key
    if not checkTokenValidHandler(key):
        return page_not_found(request)

    temp_quorpo = request.args.get('q')
    temp_id = request.args.get('id')    
    action_conn = sqlite3.connect(ASP_FOLDER+'/databases/action.db')
    action_conn.row_factory = dict_factory
    action_cur = action_conn.cursor()
    query = "DELETE FROM requests WHERE key=? AND id=? AND concod=? AND pointguyid=?;"

    if request.is_json:
        data = request.get_json()
        print(data)
        for x in data:
            if temp_quorpo == "po":
                action_cur.execute("INSERT INTO orders VALUES(NULL,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) """, (x['linkid'], x['partnumber'], x['quantity'], x['brand'], x['designation'], x['isactive'], x['aord'], x['pointguyid'], x['quorpo'], x['key'], x['requestid'], x['concod'],x['price'], x['delivery']))
                to_filter = [x['key'], x['id'], x['concod'], x['pointguyid']]
                action_cur.execute(query,to_filter)
            elif temp_quorpo == "qu":    
                action_cur.execute("INSERT INTO quotations VALUES(NULL,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) """, (x['linkid'], x['partnumber'], x['quantity'], x['brand'], x['designation'], x['isactive'], x['aord'], x['pointguyid'], x['quorpo'], x['key'], x['requestid'], x['concod'],x['price'], x['delivery']))
                to_filter = [x['aord'], x['id'], x['concod'], x['pointguyid']]
                action_cur.execute(query,to_filter)
            else:
                return page_not_found(request)
            temp_concod = x['concod']
            temp_pointguyid = x['pointguyid']
            to_filter.clear()    
        action_conn.commit()
        action_cur.close()
        action_conn.close()
        query2 = "UPDATE busreq SET done=? WHERE concode=? AND pointguyid=?;"
        to_filter2 = ['1', temp_concod, temp_pointguyid]
        base_conn = sqlite3.connect(ASP_FOLDER+'/databases/base.db')
        base_conn.row_factory = dict_factory
        base_cur = base_conn.cursor()        
        base_cur.execute(query2, to_filter2)
        base_cur.close()
        base_conn.commit()
        base_conn.close()
        
        return '''{}'''.format('Submitted!'), 200
    else:
        return page_not_found(request)  



'''
This Function is used to update all quotation items from pre-submitted items.
'''
@manager_api.route('/quo/submit', methods=['POST'])
def api_serv_edit_form(key):
    g.key = key
    if not checkTokenValid(key):
        return page_not_found(request)

    
    temp_key = request.args.get('k')    
    temp_quorpo = temp_key[-2] + temp_key[-1]
    temp_id = request.args.get('id')    

    conn_1 = sqlite3.connect(ASP_FOLDER+'/databases/action.db')
    conn_1.row_factory = dict_factory
    cur1 = conn_1.cursor()
    if request.is_json:
        data = request.get_json()
        for x in data:

            if temp_quorpo == "po":
                cur1.execute("INSERT INTO orders VALUES(NULL,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) """, (x['linkid'], x['partnumber'], x['quantity'], x['brand'], x['designation'], x['isactive'], x['aord'], x['pointguyid'], x['quorpo'], x['key'], x['requestid'], x['concod'],x['price']))
                query = "DELETE FROM requests WHERE aord=? AND price=? AND partnumber=? AND concod=? AND pointguyid=?;"
                to_filter = []
                temp_concod = x['concod']
                temp_pointguyid = x['pointguyid']
                to_filter = [x['aord'], x['price'], x['partnumber'], x['concod'], x['pointguyid']]
                cur1.execute(query,to_filter)
                conn_1.commit()
                conn_1.close()
            elif temp_quorpo == "qu":    
                cur1.execute("INSERT INTO quotations VALUES(NULL,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) """, (x['linkid'], x['partnumber'], x['quantity'], x['brand'], x['designation'], x['isactive'], x['aord'], x['pointguyid'], x['quorpo'], x['key'], x['requestid'], x['concod'],x['price']))
                conn_1.commit()
                query = "DELETE FROM requests WHERE aord=? AND price=? AND partnumber=? AND concod=? AND pointguyid=?;"
                to_filter = []
                temp_concod = x['concod']
                temp_pointguyid = x['pointguyid']
                to_filter = [x['aord'], x['price'], x['partnumber'], x['concod'], x['pointguyid']]
                cur1.execute(query,to_filter)
                conn_1.commit()
                conn_1.close()
            else:
                return page_not_found(request)

        query2 = "UPDATE busreq SET done=? WHERE concode=? AND pointguyid=?;"
        to_filter2 = ['1', temp_concod, temp_pointguyid]
        conn = sqlite3.connect(ASP_FOLDER+'/databases/base.db')
        conn.row_factory = dict_factory
        cur = conn.cursor()        
        cur.execute(query2, to_filter2)
        conn.commit()
        cur.close()
        conn.close()
        
        return '''{}'''.format('Submitted!'), 200
    else:
        return page_not_found(request)  



'''
Fetch general information about handler
Name, Phonenumber, id, customer count, etc..
'''
@manager_api.route('/general/info', methods=['GET'])
def api_serv_handler_info(key):
    g.key = key
    if not checkTokenValid(key):
        return page_not_found(request)
    cust_count = 0    
    info = {}
    temp_id = request.args.get('id')    
    
    base_conn = sqlite3.connect(ASP_FOLDER+'/databases/base.db')
    conn.row_factory = dict_factory
    cur = conn.cursor()        
    query_business = """SELECT id FROM businesses WHERE pointguyid=?"""
    query_users = """SELECT id FROM users WHERE pointguyid=?"""
    res = cur.execute(query_business,[temp_id]).fetchall()
    if len(res) > 0:
        for i in res[0]:
            cust_count+=1

    res2 = cur.execute(query_users, [temp_id]).fetchall()
    if len(res2) > 0:
        for i in res2[0]:
            cust_count+=1

    info.update({"customer_count" : str(cust_count)})
    return jsonify(info)

'''
Submit a quotation from handler to master to evaluate and price it
'''
@manager_api.route('/to/master/quo', methods=['POST'])
def api_submit_quo_master(key):
    g.key = key


'''
Fetch all quotations from master to handler
'''
@manager_api.route('/from/master/quo', methods=['GET'])
def api_serv_quo_master(key):
    g.key = key


@manager_api.route('/originality', methods=['POST'])
def api_check_originality(key):
    g.key = key






@manager_api.route('/crt/pr', methods=['GET'])
def api_create_price_request(key):
    g.key = key
    if not checkTokenValidHandler(key):
        return page_not_found(request)
    '''
        main_data = {"toCompany":"Rand Jaber", 
        "data":[ ,["BS2-2210", "BS2-2210-2RS/VT143", "25","EACH","SKF","45","1125","Immediately"],
                                                ["3310", "3310 A/C3", "50","EACH","SKF","50","2500","Immediately"],["23156 K", "23156 CCK/W33", "225","EACH","SKF","1700","382500","1-3 Weeks"]    ]}
    

    e.g:
    ["NCF 2207 ", "NCF 2207", "10","EACH","SKF","150","1500","Immediately"]


    '''
    try:
        main_data = request.get_json()

        book = openpyxl.load_workbook(ASP_FOLDER+'/templates/price_offer.xlsx')
        active_sheet = book.active
        sheet = book.get_sheet_by_name("Offer")
        ref = open(ASP_FOLDER+'/templates/sequence').read().replace("\n","")
        price_ref = "%s-PPO-2020"%ref
        sheet.cell(row=2, column=1).value = "PR Reference : %s"%price_ref
        sheet.cell(row=3, column=1).value = "Company : %s"%main_data['customername']
        sheet.cell(row=2, column=7).value = "Request Date : %s"%str(time.strftime("%x"))
        sheet.cell(row=3, column=7).value = "Offer Date : %s"%str(time.strftime("%x"))
        current_row = 10
        total_pr = 0
        itemscount = len(main_data['items'])-1
        sheet.move_range("A11:I18", rows=itemscount)
        ft1 = Font(name='Arial', size=14)
        al1=Alignment(horizontal='center', vertical='center')
        new_merged_cells = []
        for merged_cells in sheet.merged_cell_ranges[-4:]:
            try:
                sheet.unmerge_cells(merged_cells.coord)
                new_merged_cells.append(merged_cells.coord)
            except:
                print("...")
        mydata = ReArrangeData(main_data['items'])
        for j,n in enumerate(mydata):
            
            sheet.cell(row=current_row, column=1).value = str(j+1)
            sheet.cell(row=current_row, column=1).font = ft1
            sheet.cell(row=current_row, column=1).alignment = al1
            for i in range(2,10):
                sheet.cell(row=current_row, column=i).value = str(n[i-2])
                sheet.cell(row=current_row,column=i).font = ft1
                sheet.cell(row=current_row,column=i).alignment = al1
            total_pr += int(n[6])
            current_row += 1
        for tb_merge_cells in new_merged_cells:
            idxs = re.findall(r'\d+', tb_merge_cells)
            for h in idxs:
                tb_merge_cells = tb_merge_cells.replace(h, str(int(h)+1))
            try:
                sheet.merge_cells(tb_merge_cells)
            except:
                print("...")
        sheet.cell(row=current_row, column=8).value = str(round(total_pr, 2))
        sheet.merge_cells("C%d:G%d"%(current_row,current_row))
        sheet.cell(row=current_row+1, column=8).value = str(round(0.16*total_pr, 2))
        sheet.merge_cells("C%d:G%d"%(current_row+1,current_row+1))
        sheet.cell(row=current_row+2, column=8).value = str(round(1.16* total_pr, 2)) 
        sheet.merge_cells("C%d:G%d"%(current_row+2,current_row+2))
        sheet.cell(row=8,column=1).alignment = Alignment(wrap_text=False)
        book.save(OFFER_DIRECTORIES+price_ref+".xlsx")
        open(ASP_FOLDER+'/templates/sequence', 'w').write("%s"%str(int(ref)+1))
        return "{}".format(price_ref)
    except:
        return page_not_found(request)








